from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def read_workbook(path: str) -> tuple[list[dict], list[dict]]:
    source = Path(path)
    if not source.exists():
        return [], []

    workbook = load_workbook(source, data_only=True, read_only=True, keep_vba=True)
    addresses: list[dict] = []
    for number, address, phone in workbook["DATABASE"].iter_rows(min_row=2, values_only=True):
        if not number or not address:
            continue
        addresses.append(
            {"number": int(number), "address": text(address), "phone": text(phone)}
        )

    address_lookup = {item["address"]: item["number"] for item in addresses}
    fallback_date = date(2026, 8, 28)
    now = datetime.now(timezone.utc)
    shipments: list[dict] = []
    for row in workbook["GRADLE"].iter_rows(min_row=2, values_only=True):
        code, bag, sender, sender_address, recipient, recipient_address, phone, contents, attendant = row[:9]
        if not code or not bag:
            continue
        clean_recipient_address = text(recipient_address)
        shipments.append(
            {
                "code": "".join(text(code).upper().split()),
                "bag_number": int(bag),
                "shipper_name": text(sender),
                "shipper_address": text(sender_address),
                "consignee_name": text(recipient),
                "address_number": address_lookup.get(clean_recipient_address, 1),
                "consignee_address": clean_recipient_address,
                "phone": text(phone),
                "contents": text(contents).upper(),
                "attendant": text(attendant) or "SIN ASIGNAR",
                "shipment_date": fallback_date.isoformat(),
                "status": "registrado",
                "customs_type": "UNSOLICITED",
                "quantity": 2,
                "created_at": now,
                "updated_at": now,
            }
        )
    workbook.close()
    return addresses, shipments


async def seed_database(database, workbook_path: str) -> dict[str, int]:
    addresses, shipments = read_workbook(workbook_path)
    inserted_addresses = 0
    inserted_shipments = 0

    if addresses and await database.addresses.count_documents({}) == 0:
        result = await database.addresses.insert_many(addresses)
        inserted_addresses = len(result.inserted_ids)

    if shipments and await database.shipments.count_documents({}) == 0:
        result = await database.shipments.insert_many(shipments)
        inserted_shipments = len(result.inserted_ids)

    return {"addresses": inserted_addresses, "shipments": inserted_shipments}

